from lector_excel import Lector
from openpyxl import load_workbook
import os
 
# Funcion con parametro de entrada
def leer_archivo(nombre_archivo, col):

    # Crear un objeto para leer a partir de la clase
    archivo_excel = Lector(nombre_archivo)

    # Leer columna
    columna = archivo_excel.leer_columna(col)
    
    # Eliminar primer elemento por ser string
    columna.pop(0)
    
    # Sumar los valores de la columna
    valColumna = [celda for celda in columna]
    return valColumna
 
carpeta = "archivos"
# listar todos los archivos de una carpeta
lista_archivos = os.listdir(carpeta)

sumas_total_v = []
sumas_total_c = []
sumas_total_g = []
mes_ventaAlta = ""
promedios_c = 0
promedios_g = 0
mes=0

# Recorriendo la lista de archivos
for nombre_archivo in lista_archivos:
    #_____________________________
    #_______________________________

    venta = leer_archivo((f'archivos//{nombre_archivo}'), 'B')
    costo = leer_archivo((f'archivos//{nombre_archivo}'), 'C')
    ganancia = leer_archivo((f'archivos//{nombre_archivo}'), 'D')
    
    promedio_costo = round(sum(costo) / len(costo),1)
    promedio_ganancia = round(sum(ganancia) / len(ganancia),1)
    
    promedios_c += promedio_costo
    promedios_g += promedio_ganancia
    mes+=1

    for v in venta:
        sumas_total_v.append(v)

    for c in costo:
        sumas_total_c.append(c)

    for g in ganancia:
        sumas_total_g.append(g)

    print()
    nombre_mes = nombre_archivo[3:-5]
    print("---------  ", nombre_mes.upper(),"  ---------")
    print("Suma de costos: \t", sum(costo))
    print("Promedio de costos: \t", promedio_costo)
    print("Suma de Ganancias: \t", sum(ganancia))
    print("Promedio de Ganancias: \t", promedio_ganancia)
    print("-------------------------------")


venta_mayor = max(sumas_total_v)
costo_mayor = max(sumas_total_c)
print()
print("------  TODOS LOS MESES   ------")
print("Total de costos: \t",sum(sumas_total_c))
print("Promedio de costos: \t",promedios_c/mes)
print("Total de ganacias: \t",sum(sumas_total_g))
print("Promedio de ganancias: \t",promedios_g/mes)
print("-------------------------------")


# listar todos los archivos de una carpeta

mesV = [""]
for mes in lista_archivos:


    print(mes)
    
    wb = load_workbook((f'archivos//{mes}')) 
    ws = wb["Sheet1"]
    ws = wb.active   
    def buscar_Valor(searchVal,columna):
        for i in range(1, ws.max_row + 1):
            for j in range(columna, ws.max_column + 1):
                if j > columna:
                    break
                if searchVal == ws.cell(i,j).value:
                    mesV.insert(0,mes)

    buscar_Valor(venta_mayor,2)
mesVentas = mesV[0]
print()
mes_ventasMax = mesVentas[3:-5]
print(mes_ventasMax.upper() , "MES con venta más alta: ",max(sumas_total_v))

mesC = [""]
for mes in lista_archivos:
    wb = load_workbook(mes) 
    ws = wb.active   
    
    def buscar_Costo(searchVal,columna):
        for i in range(1, ws.max_row + 1):
            for j in range(columna, ws.max_column + 1):
                if j > columna:
                    break
                if searchVal == ws.cell(i,j).value:
                    mesC.insert(0,mes)

    buscar_Costo(costo_mayor,3)

mesCosto = mesC[0]
mes_costoMax = mesCosto[3:-5]


print(mes_costoMax.upper() ,"MES con costo más alto: \t",max(sumas_total_c))
print("MES con ganancia más baja: \t",min(sumas_total_g))
print("-------------------------------")
print(venta_mayor)
print(costo_mayor)