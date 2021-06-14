from openpyxl import load_workbook
import re

# open workbook 
excel_file1 = '01-enero.xlsx'
wb1 = load_workbook(excel_file1) # wrbk
ws1 = wb1["Sheet1"]

excel_file2 = '02-febrero.xlsx'
wb2 = load_workbook(excel_file2) # book_1
ws2 = wb2["Sheet1"] # sh_1

# fn to search all sheets in workbook
def myfind(wb,s):
    for ws in wb.worksheets:       
        for c in range(5,ws.max_column+1):
            for r in range(1,ws.max_row+1):
                txt = ws.cell(r,c).value 
                if txt is None:
                    pass
                elif re.search(s,txt):
                    print("Found",s,txt,ws,r,c)

# scan col C
for r in range(1,ws2.max_row+1):
    s = ws1.cell(r, 3).value 
    if s is None: 
        pass
    else: 
        print(r,s)
        myfind(wb1,s)