from openpyxl import load_workbook
import re


excel_file2 = 'Book2_test.xlsx'
wb2 = load_workbook(excel_file2) # book_1
ws2 = wb2["Sheet1"] # sh_1


ws = wb2.active    
maxValue=9643

def wordfinder(searchData):
    for i in range(1, ws.max_row + 1):
        
        for j in range(2, ws.max_column + 1):
            if j != 2:
                pass
            elif searchData == ws.cell(i,j).value:
                print("Encontrado: ",(ws.cell(i,j)).value )
                print(ws.cell(i,j))
                print(wb2.sheetnames) 




wordfinder(maxValue)

listamayor=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26]
print(listamayor)


for x in listamayor:
    if 263 == x:
        print("Encontre al 26")


